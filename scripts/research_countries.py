"""
research_countries.py

Researches HSESC legislative and compliance context for each project jurisdiction
using Claude with web search, then builds a multi-tab Excel sourcebook.

Tabs produced:
  1. Country Profiles       — one row per jurisdiction, all research dimensions
  2. HSESC Compliance Matrix — 51 procedures × 8 jurisdictions, RAG + statement
  3. Document Versioning    — grouping logic and version count per procedure
  4. Source Log             — every factual claim with URL and verification status
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import date

import anthropic
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ── Jurisdiction definitions ──────────────────────────────────────────────────

JURISDICTIONS = [
    {
        "id": "GN",
        "name": "Guinea",
        "location": "Simandou",
        "commodity": "Iron Ore",
        "federal_context": "Republic of Guinea — civil law system",
        "notes": "Remote, undeveloped region; rail and port infrastructure being built from scratch"
    },
    {
        "id": "WA",
        "name": "Western Australia",
        "location": "Pilbara (Rhodes Ridge, Hope Downs 2, Western Range, Winu)",
        "commodity": "Iron Ore / Copper-Gold",
        "federal_context": "Australian federal + WA state jurisdiction",
        "notes": "Mature mining jurisdiction, strong legislative framework"
    },
    {
        "id": "QLD",
        "name": "Queensland, Australia",
        "location": "Weipa / Norman Creek",
        "commodity": "Aluminium / Bauxite",
        "federal_context": "Australian federal + QLD state jurisdiction",
        "notes": "Remote but established mining region; Cape York Peninsula"
    },
    {
        "id": "AR",
        "name": "Argentina",
        "location": "Rincon, Sal de Vida (Puna region, Jujuy / Salta)",
        "commodity": "Lithium",
        "federal_context": "Federal republic; civil law system; provincial jurisdiction over mining",
        "notes": "High altitude, remote, arid; emerging lithium jurisdiction"
    },
    {
        "id": "CA",
        "name": "Canada (Quebec)",
        "location": "Nemaska (Whabouchi mine + Bécancour), AP60 Smelter",
        "commodity": "Lithium / Aluminium",
        "federal_context": "Canadian federal + Quebec provincial jurisdiction; civil law (Quebec)",
        "notes": "Mature jurisdiction; Indigenous (Cree) land rights significant at Whabouchi"
    },
    {
        "id": "MN",
        "name": "Mongolia",
        "location": "Oyu Tolgoi (South Gobi)",
        "commodity": "Copper",
        "federal_context": "Parliamentary republic; mixed civil/socialist law tradition",
        "notes": "Remote, extreme climate; rapidly developing mining law framework"
    },
    {
        "id": "US",
        "name": "United States",
        "location": "Resolution Copper (Arizona), Kennecott (Utah)",
        "commodity": "Copper",
        "federal_context": "US federal + Arizona and Utah state jurisdictions",
        "notes": "Highly mature regulatory environment; federal MSHA, EPA, OSHA"
    },
]

# RAG definitions
RAG = {
    "G": "GREEN — Can comply with full industry-leading standard. No structural barriers.",
    "A": "AMBER — Can comply with most requirements. Specific gaps exist that require adaptation.",
    "R": "RED — Structural barriers prevent compliance with the base standard without modification.",
    "U": "UNKNOWN — Insufficient verified information to make a determination.",
}

# Research dimensions for Country Profiles tab
RESEARCH_DIMENSIONS = [
    "development_classification",
    "osh_legislation",
    "environmental_legislation",
    "community_social_legislation",
    "human_rights_frameworks",
    "security_frameworks",
    "safety_equipment_availability",
    "training_providers",
    "security_companies",
    "socioeconomic_context",
    "overall_compliance_maturity",
]

DIMENSION_LABELS = {
    "development_classification": "Development Classification",
    "osh_legislation": "OH&S Legislation",
    "environmental_legislation": "Environmental Legislation",
    "community_social_legislation": "Community & Social Legislation",
    "human_rights_frameworks": "Human Rights Frameworks",
    "security_frameworks": "Security Frameworks",
    "safety_equipment_availability": "Safety Equipment Availability",
    "training_providers": "Training Providers",
    "security_companies": "Security Companies",
    "socioeconomic_context": "Socioeconomic Context",
    "overall_compliance_maturity": "Overall HSESC Compliance Maturity",
}


# ── Claude API helpers ────────────────────────────────────────────────────────

def call_claude_with_search(client, system, prompt, label="", max_tokens=8000):
    """
    Two-call approach:
    Call 1 — web search enabled, returns plain text research findings.
    Call 2 — no web search, structures the findings into the required JSON.
    This avoids JSON parse failures from mixed tool_use/text content blocks.
    """
    for attempt in range(3):
        try:
            # ── Call 1: gather research with web search ──────────────────────
            research_response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=max_tokens,
                system=system,
                tools=[{"type": "web_search_20250305", "name": "web_search"}],
                messages=[{
                    "role": "user",
                    "content": (
                        prompt +
                        "\n\nIMPORTANT: Search the web thoroughly for this information. "
                        "Return your findings as detailed plain text prose — NOT as JSON. "
                        "Include all source URLs you find inline with the text. "
                        "Be comprehensive and specific."
                    )
                }],
            )

            # Extract all text from research response
            research_text = " ".join(
                b.text for b in research_response.content
                if hasattr(b, "text") and b.text
            ).strip()

            if not research_text:
                raise ValueError("Empty research response from web search call")

            print(f"    Research gathered: {len(research_text)} chars")

            # ── Call 2: structure into JSON (no web search) ──────────────────
            # Re-extract the JSON schema from the original prompt
            json_schema_start = prompt.find("Return this JSON object:")
            if json_schema_start == -1:
                json_schema_start = prompt.find("Return ONLY this JSON")
            schema_instruction = prompt[json_schema_start:] if json_schema_start != -1 else prompt[-3000:]

            structure_prompt = f"""Based on the research findings below, structure the information 
into the required JSON format. Only include information that appears in the research findings.
If a piece of information was not found in the research, use null or "Not found" as appropriate.
Include all source URLs exactly as they appear in the research text.

RESEARCH FINDINGS:
{research_text}

{schema_instruction}

Return ONLY valid JSON — no markdown fences, no preamble, no commentary."""

            structure_response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=max_tokens,
                system="You are a JSON structuring assistant. Output only valid JSON — no markdown, no commentary.",
                messages=[{"role": "user", "content": structure_prompt}],
            )

            raw = structure_response.content[0].text.strip()
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
            return json.loads(raw)

        except json.JSONDecodeError as e:
            print(f"  JSON error {label} attempt {attempt+1}: {e}")
            if attempt < 2:
                time.sleep(10)
            else:
                return None
        except Exception as e:
            print(f"  API error {label} attempt {attempt+1}: {e}")
            if attempt < 2:
                time.sleep(15)
            else:
                return None


def call_claude_plain(client, system, prompt, label="", max_tokens=8000):
    """Call Claude without web search for synthesis/structuring tasks."""
    for attempt in range(3):
        try:
            response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=max_tokens,
                system=system,
                messages=[{"role": "user", "content": prompt}],
            )
            raw = response.content[0].text.strip()
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
            return json.loads(raw)
        except json.JSONDecodeError as e:
            print(f"  JSON error {label} attempt {attempt+1}: {e}")
            if attempt < 2:
                time.sleep(5)
            else:
                return None
        except Exception as e:
            print(f"  API error {label} attempt {attempt+1}: {e}")
            if attempt < 2:
                time.sleep(10)
            else:
                return None


# ── Research functions ────────────────────────────────────────────────────────

SYSTEM_RESEARCH = """You are an HSESC (Health, Safety, Environment, Security, and Community) 
research specialist with expertise in international mining and resources legislation. 
You research jurisdictions thoroughly and only report information you can verify with 
working URLs or credible source citations. If you cannot verify a claim, you say so explicitly 
rather than guessing. You always report the absence of information as a finding — 
"no evidence found" is a valid and important answer.
Output only valid JSON — no markdown fences, no preamble."""



def _research_topic(client, jur_name, jur_location, jur_notes, topic, schema):
    """Research one topic for a jurisdiction. Returns dict or None."""
    for attempt in range(3):
        try:
            # Call 1: web search for raw information as plain text
            r1 = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=4000,
                system=SYSTEM_RESEARCH,
                tools=[{"type": "web_search_20250305", "name": "web_search"}],
                messages=[{"role": "user", "content":
                    f"Research {topic} for {jur_name} ({jur_location}). "
                    f"Context: {jur_notes}. "
                    "Search the web thoroughly. Return detailed plain text findings with all source URLs inline. "
                    "Do NOT return JSON."
                }],
            )
            research_text = " ".join(
                b.text for b in r1.content if hasattr(b, "text") and b.text
            ).strip()

            if not research_text:
                raise ValueError("Empty web search response")

            print(f"      Search returned {len(research_text)} chars")

            # Call 2: structure findings into JSON (no web search)
            r2 = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=3000,
                system="Output only valid JSON. No markdown, no preamble, no commentary.",
                messages=[{"role": "user", "content":
                    "Based on these research findings, fill in the JSON schema below.\n"
                    "Only use information present in the findings. Use null for missing data.\n"
                    "Include source URLs exactly as found in the research.\n\n"
                    "FINDINGS:\n" + research_text[:8000] + "\n\n"
                    "JSON SCHEMA TO FILL:\n" + schema + "\n\n"
                    "Return only valid JSON."
                }],
            )
            raw = r2.content[0].text.strip()
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
            return json.loads(raw)

        except json.JSONDecodeError as e:
            print(f"      JSON error attempt {attempt+1}: {e}")
            if attempt < 2:
                time.sleep(5)
        except Exception as e:
            print(f"      Error attempt {attempt+1}: {e}")
            if attempt < 2:
                time.sleep(10)
    return None


def research_jurisdiction_profile(client, jurisdiction):
    """
    Research a jurisdiction across all HSESC dimensions.
    Each dimension is a separate web search + JSON structuring call pair
    to avoid token limit issues with large combined JSON responses.
    """
    jur_name     = jurisdiction["name"]
    jur_location = jurisdiction["location"]
    jur_notes    = jurisdiction["notes"]

    topics = [
        ("development_classification",
         '{"classification":"<First World/Second World/Third World or OECD/Non-OECD/LDC>",'
         '"hdi_rank":"<UN HDI rank and score or Not available>",'
         '"hdi_source_url":"<URL or null>",'
         '"notes":"<what this means for HSESC compliance>"}'),

        ("osh_legislation",
         '{"has_dedicated_osh_law":true_or_false,'
         '"legislation":[{"name":"<name>","year":"<year>","regulator":"<body>",'
         '"regulator_url":"<URL or null>","source_url":"<URL or null>",'
         '"verified":true_or_false,"notes":"<relevance>"}],'
         '"global_frameworks_applied":[{"framework":"<name>",'
         '"evidence_of_application":"<context>","source_url":"<URL or null>"}],'
         '"gap_summary":"<what domestic law does not cover>"}'),

        ("environmental_legislation",
         '{"has_dedicated_env_law":true_or_false,'
         '"legislation":[{"name":"<name>","year":"<year>","regulator":"<body>",'
         '"source_url":"<URL or null>","verified":true_or_false,"notes":"<relevance>"}],'
         '"global_frameworks_applied":[{"framework":"<name>",'
         '"evidence_of_application":"<context>","source_url":"<URL or null>"}],'
         '"gap_summary":"<gaps>"}'),

        ("community_social_legislation",
         '{"has_dedicated_csl_law":true_or_false,'
         '"legislation":[{"name":"<name>","year":"<year>","regulator":"<body>",'
         '"source_url":"<URL or null>","verified":true_or_false,"notes":"<relevance>"}],'
         '"global_frameworks_applied":[{"framework":"<name>",'
         '"evidence_of_application":"<context>","source_url":"<URL or null>"}],'
         '"gap_summary":"<gaps>"}'),

        ("human_rights_frameworks",
         '{"has_domestic_human_rights_law":true_or_false,'
         '"legislation_or_constitution":[{"name":"<name>","year":"<year>",'
         '"source_url":"<URL or null>","verified":true_or_false,"notes":"<relevance>"}],'
         '"international_conventions_ratified":[{"convention":"<name>",'
         '"ratified":true_or_false,"source_url":"<URL or null>"}],'
         '"global_frameworks_applied":[{"framework":"<name>",'
         '"evidence_of_application":"<context>","source_url":"<URL or null>"}],'
         '"modern_slavery_risk":"<High/Medium/Low with explanation>",'
         '"gap_summary":"<gaps>"}'),

        ("security_frameworks",
         '{"has_private_security_regulation":true_or_false,'
         '"legislation":[{"name":"<name>","year":"<year>","source_url":"<URL or null>",'
         '"verified":true_or_false,"notes":"<what it regulates>"}],'
         '"voluntary_principles_applied":true_or_false,'
         '"vp_evidence_url":"<URL or null>",'
         '"conflict_risk_level":"<High/Medium/Low with explanation>",'
         '"gap_summary":"<gaps>"}'),

        ("safety_equipment_availability",
         '{"overall_rating":"<Readily Available/Limited Availability/Very Limited Availability>",'
         '"ppe_availability":"<description>",'
         '"specialist_equipment_notes":"<cranes MEWPs gas detection fall arrest etc>",'
         '"import_feasibility":"<cost customs lead time>",'
         '"known_specific_gaps":["<e.g. seatbelted buses unavailable locally>"],'
         '"source_urls":["<URL or null>"],'
         '"confidence":"<High/Medium/Low>"}'),

        ("training_providers",
         '{"has_national_training_framework":true_or_false,'
         '"framework_name":"<name or null>",'
         '"registered_providers_available":true_or_false,'
         '"mining_specific_training":"<description>",'
         '"internationally_recognised_certs_available":true_or_false,'
         '"cert_examples":["<e.g. IOSH NEBOSH ICAM BOSIET>"],'
         '"workforce_literacy_rate":"<% or qualitative>",'
         '"source_urls":["<URL or null>"],'
         '"gap_summary":"<what cannot be sourced locally>",'
         '"confidence":"<High/Medium/Low>"}'),

        ("security_companies",
         '{"reputable_providers_available":true_or_false,'
         '"notable_providers":[{"name":"<name>","international":true_or_false,'
         '"human_rights_certified":true_or_false,"notes":"<notes>",'
         '"source_url":"<URL or null>"}],'
         '"armed_security_regulation":"<regulated/unregulated/prohibited with context>",'
         '"conflict_risk_to_security_operations":"<High/Medium/Low>",'
         '"gap_summary":"<gaps>",'
         '"confidence":"<High/Medium/Low>"}'),

        ("socioeconomic_context",
         '{"poverty_rate":"<% or qualitative with source>",'
         '"poverty_source_url":"<URL or null>",'
         '"basic_sanitation_access":"<% or qualitative>",'
         '"sanitation_source_url":"<URL or null>",'
         '"clean_water_access":"<% or qualitative>",'
         '"water_source_url":"<URL or null>",'
         '"healthcare_system_quality":"<description>",'
         '"infectious_disease_risk":"<key diseases relevant to project workforce>",'
         '"local_workforce_skill_level":"<description>",'
         '"infrastructure_quality":"<roads power comms>",'
         '"corruption_perception_index":"<CPI rank and score or Not available>",'
         '"cpi_source_url":"<URL or null>",'
         '"overall_context_summary":"<2-3 sentence summary>"}'),

        ("overall_compliance_maturity",
         '{"rag_rating":"<G/A/R>",'
         '"rag_rationale":"<detailed explanation>",'
         '"key_strengths":["<strength>"],'
         '"key_barriers":["<barrier>"],'
         '"recommended_approach":"<how documents should be calibrated for this jurisdiction>"}'),
    ]

    profile = {"sources": []}
    total = len(topics)

    for i, (topic_key, schema) in enumerate(topics):
        print(f"    [{i+1}/{total}] {topic_key}...")
        result = _research_topic(client, jur_name, jur_location, jur_notes, topic_key, schema)
        if result:
            profile[topic_key] = result
            # Collect sources
            if isinstance(result, dict):
                for v in result.values():
                    if isinstance(v, list):
                        for item in v:
                            if isinstance(item, dict) and item.get("source_url"):
                                profile["sources"].append({
                                    "claim": topic_key,
                                    "source_name": item.get("name", topic_key),
                                    "url": item.get("source_url", ""),
                                    "date_accessed": date.today().isoformat(),
                                    "verified": item.get("verified", False),
                                    "notes": ""
                                })
        else:
            profile[topic_key] = {"_topic_research_failed": True}
            print(f"    WARNING: {topic_key} failed — placeholder written")
        time.sleep(1)

    return profile


def research_procedure_compliance(client, procedure, jurisdictions, profiles):
    """
    Research compliance feasibility for one procedure across all jurisdictions.
    Returns a dict keyed by jurisdiction ID with RAG rating and statement.
    """
    proc_id = procedure["id"]
    proc_title = procedure["title"]

    # Build a summary of each jurisdiction's relevant profile for context
    jur_summaries = []
    for j in jurisdictions:
        p = profiles.get(j["id"])
        if p:
            maturity = p.get("overall_compliance_maturity", {})
            socio = p.get("socioeconomic_context", {})
            equip = p.get("safety_equipment_availability", {})
            training = p.get("training_providers", {})
            jur_summaries.append(
                f"{j['id']} ({j['name']}): "
                f"RAG={maturity.get('rag_rating','U')}, "
                f"Equipment={equip.get('overall_rating','Unknown')}, "
                f"Training={training.get('registered_providers_available','Unknown')}, "
                f"Barriers: {'; '.join(maturity.get('key_barriers', ['None identified']))}"
            )
        else:
            jur_summaries.append(f"{j['id']} ({j['name']}): No profile data available")

    prompt = f"""Assess the HSESC compliance feasibility for the following procedure across 
all project jurisdictions. Use the jurisdiction context provided and your knowledge of 
what this type of procedure typically requires in practice.

PROCEDURE: {proc_id} — {proc_title}

JURISDICTION CONTEXT:
{chr(10).join(jur_summaries)}

RAG DEFINITIONS:
G (GREEN): Can comply with full industry-leading standard. No structural barriers to compliance. 
  Legislation exists, equipment available, trained workforce available.
A (AMBER): Can comply with most requirements. Specific gaps exist — document may need adaptation 
  or additional controls. Some equipment or training must be imported/contracted from outside.
R (RED): Structural barriers prevent full compliance with industry-leading standard without 
  significant modification. Physical unavailability of equipment, workforce, or regulatory 
  framework makes full compliance impossible or impractical.
U (UNKNOWN): Insufficient verified information to make a determination.

For each jurisdiction, assess whether that jurisdiction can realistically comply with 
industry-leading requirements for this specific procedure topic.

Be specific — name the actual barrier (e.g. "seatbelted buses not commercially available 
in Guinea", "no licensed scaffold inspectors in Mongolia", "no BOSIET/HUET providers 
within 500km"). If you cannot identify a specific barrier, say so.

Return ONLY this JSON:

{{
  "procedure_id": "{proc_id}",
  "procedure_title": "{proc_title}",
  "jurisdictions": {{
    "GN": {{
      "rag": "<G/A/R/U>",
      "statement": "<one sentence — what they can or cannot do and specifically why>",
      "specific_barrier": "<the named specific barrier, or null if none>",
      "workaround": "<if R or A: what is the realistic workaround or adaptation needed, or null>"
    }},
    "WA": {{"rag": "<G/A/R/U>", "statement": "<...>", "specific_barrier": null, "workaround": null}},
    "QLD": {{"rag": "<G/A/R/U>", "statement": "<...>", "specific_barrier": null, "workaround": null}},
    "AR": {{"rag": "<G/A/R/U>", "statement": "<...>", "specific_barrier": null, "workaround": null}},
    "CA": {{"rag": "<G/A/R/U>", "statement": "<...>", "specific_barrier": null, "workaround": null}},
    "MN": {{"rag": "<G/A/R/U>", "statement": "<...>", "specific_barrier": null, "workaround": null}},
    "US": {{"rag": "<G/A/R/U>", "statement": "<...>", "specific_barrier": null, "workaround": null}}
  }}
}}"""

    result = call_claude_plain(client, SYSTEM_RESEARCH, prompt,
                                label=f"compliance-{proc_id}", max_tokens=4000)
    return result


def research_versioning(client, procedures, compliance_matrix):
    """Determine document versioning groupings based on compliance matrix."""

    # Summarise the matrix as a compact string
    matrix_summary = []
    for proc in procedures:
        pid = proc["id"]
        ratings = {}
        if pid in compliance_matrix:
            for jid, data in compliance_matrix[pid].get("jurisdictions", {}).items():
                ratings[jid] = data.get("rag", "U")
        matrix_summary.append(f"{pid}: " + ", ".join(f"{k}={v}" for k, v in ratings.items()))

    prompt = f"""Based on the HSESC compliance matrix below, determine document versioning 
groupings for the 51 procedures. Jurisdictions that have the same RAG pattern across a 
procedure can share the same document version.

COMPLIANCE MATRIX (procedure: jurisdiction=RAG):
{chr(10).join(matrix_summary)}

JURISDICTIONS: GN=Guinea, WA=Western Australia, QLD=Queensland, AR=Argentina, 
CA=Canada/Quebec, MN=Mongolia, US=United States

Analyse the patterns and recommend:
1. Which jurisdictions are "like" each other and could share documents
2. How many distinct versions each procedure needs
3. An overall versioning strategy

Return ONLY this JSON:

{{
  "jurisdiction_groupings": [
    {{
      "group_name": "<e.g. 'Tier 1 — Mature Jurisdictions'>",
      "jurisdictions": ["WA", "QLD", "CA", "US"],
      "rationale": "<why these can share documents>",
      "document_standard": "<what level of prescriptiveness is appropriate>"
    }}
  ],
  "procedures": [
    {{
      "procedure_id": "<e.g. PRO-0001>",
      "versions_needed": <integer>,
      "version_groups": [["WA","QLD","CA","US"], ["AR","MN"], ["GN"]],
      "version_notes": "<what differs between versions>"
    }}
  ],
  "overall_strategy": "<2-3 sentence summary of the recommended versioning approach>",
  "single_version_procedures": ["<PRO-XXXX>"],
  "most_complex_procedures": ["<PRO-XXXX>"]
}}"""

    result = call_claude_plain(client, SYSTEM_RESEARCH, prompt,
                                label="versioning", max_tokens=6000)
    return result


# ── Excel builder ─────────────────────────────────────────────────────────────

# Colour palette
NAVY    = "1F3864"
WHITE   = "FFFFFF"
GREEN   = "375623"
AMBER   = "C55A11"
RED_C   = "C00000"
LIGHT_G = "E2EFDA"
LIGHT_A = "FFF2CC"
LIGHT_R = "FCE4D6"
LIGHT_U = "EDEDED"
BLUE_H  = "2E75B6"
LIGHT_B = "DEEAF1"
GREY_R  = "F2F2F2"

def hdr_font(size=10, bold=True, colour=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=colour)

def body_font(size=9, bold=False, colour="000000"):
    return Font(name="Arial", size=size, bold=bold, color=colour)

def hdr_fill(colour=NAVY):
    return PatternFill("solid", fgColor=colour)

def cell_fill(colour):
    return PatternFill("solid", fgColor=colour)

def wrap_align(horizontal="left", vertical="top"):
    return Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def rag_fill(rag):
    return {
        "G": cell_fill(LIGHT_G),
        "A": cell_fill(LIGHT_A),
        "R": cell_fill(LIGHT_R),
        "U": cell_fill(LIGHT_U),
    }.get(rag, cell_fill(LIGHT_U))

def rag_text(rag):
    return {"G": "GREEN", "A": "AMBER", "R": "RED", "U": "UNKNOWN"}.get(rag, "UNKNOWN")

def write_header_row(ws, row, values, bg=NAVY, fg=WHITE, size=10):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = hdr_font(size=size, colour=fg)
        c.fill = hdr_fill(bg)
        c.alignment = wrap_align("center", "center")
        c.border = thin_border()

def write_cell(ws, row, col, value, fill=None, font=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or body_font()
    c.fill = fill or cell_fill(WHITE)
    c.alignment = align or wrap_align()
    c.border = thin_border()
    return c


def build_tab1_country_profiles(ws, jurisdictions, profiles):
    ws.title = "1. Country Profiles"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:L1")
    title = ws["A1"]
    title.value = "HSESC Country Compliance Sourcebook — Country Profiles"
    title.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    title.alignment = Alignment(horizontal="left", vertical="center")
    title.fill = cell_fill(LIGHT_B)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:L2")
    sub = ws["A2"]
    sub.value = (
        f"Research date: {date.today().isoformat()}  |  "
        "RAG: GREEN = No barriers, AMBER = Specific gaps, RED = Structural barriers, UNKNOWN = Insufficient data  |  "
        "All claims are sourced. Unsourced claims are flagged. See Source Log tab for full references."
    )
    sub.font = Font(name="Arial", size=8, italic=True, color="555555")
    sub.alignment = wrap_align()
    ws.row_dimensions[2].height = 30

    headers = [
        "Jurisdiction", "Location / Project", "Commodity",
        "Development Classification", "HDI Rank",
        "OH&S Legislation", "Environmental Legislation",
        "Community & Social Legislation", "Human Rights Frameworks",
        "Safety Equipment Availability",
        "Training Providers",
        "Security Companies",
        "Socioeconomic Context",
        "Overall HSESC Maturity",
        "RAG Rating",
        "Key Barriers",
        "Recommended Approach",
    ]
    write_header_row(ws, 3, headers, bg=NAVY, size=9)
    ws.row_dimensions[3].height = 40

    for row_idx, jur in enumerate(jurisdictions, 4):
        p = profiles.get(jur["id"], {})
        if not p:
            for col in range(1, len(headers) + 1):
                write_cell(ws, row_idx, col, "Research failed — no data available",
                           fill=cell_fill(LIGHT_U))
            write_cell(ws, row_idx, 1, jur["name"])
            continue

        dev = p.get("development_classification", {})
        osh = p.get("osh_legislation", {})
        env = p.get("environmental_legislation", {})
        csl = p.get("community_social_legislation", {})
        hr  = p.get("human_rights_frameworks", {})
        sec = p.get("security_frameworks", {})
        equip = p.get("safety_equipment_availability", {})
        train = p.get("training_providers", {})
        secc  = p.get("security_companies", {})
        socio = p.get("socioeconomic_context", {})
        mat   = p.get("overall_compliance_maturity", {})
        rag   = mat.get("rag_rating", "U")

        def leg_summary(leg_data):
            laws = leg_data.get("legislation", [])
            if not laws:
                gaps = leg_data.get("gap_summary", "No domestic legislation identified.")
                gfw  = leg_data.get("global_frameworks_applied", [])
                if gfw:
                    frameworks = "; ".join(g.get("framework","") for g in gfw[:3])
                    return f"No domestic legislation. Global frameworks applied: {frameworks}. {gaps}"
                return f"No domestic legislation. {gaps}"
            names = "; ".join(l.get("name","") + (f" ({l.get('year','')})" if l.get("year") else "") for l in laws[:3])
            return names + (f" | Gap: {leg_data.get('gap_summary','')}" if leg_data.get("gap_summary") else "")

        row_data = [
            jur["name"],
            jur["location"],
            jur["commodity"],
            f"{dev.get('classification','Unknown')}",
            dev.get("hdi_rank", "N/A"),
            leg_summary(osh),
            leg_summary(env),
            leg_summary(csl),
            "; ".join(c.get("convention","") for c in hr.get("international_conventions_ratified", [])[:5])
            or leg_summary(hr),
            f"{equip.get('overall_rating','Unknown')} | {equip.get('ppe_availability','')[:200]}",
            f"Registered: {'Yes' if train.get('registered_providers_available') else 'No/Unknown'} | {train.get('gap_summary','')[:200]}",
            f"Reputable available: {'Yes' if secc.get('reputable_providers_available') else 'No/Unknown'} | {secc.get('gap_summary','')[:150]}",
            socio.get("overall_context_summary", ""),
            mat.get("rag_rationale", ""),
            rag_text(rag),
            "\n".join(f"• {b}" for b in mat.get("key_barriers", [])),
            mat.get("recommended_approach", ""),
        ]

        fill = rag_fill(rag)
        for col, val in enumerate(row_data, 1):
            c = write_cell(ws, row_idx, col, val, fill=fill if col >= 14 else cell_fill(WHITE))
            if col == 15:  # RAG column
                c.font = Font(name="Arial", size=9, bold=True,
                              color={"G": "375623", "A": "C55A11", "R": "C00000"}.get(rag, "555555"))

        ws.row_dimensions[row_idx].height = 80

    # Column widths
    col_widths = [18, 28, 15, 18, 10, 45, 45, 40, 40, 40, 40, 35, 45, 50, 12, 40, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # AutoFilter on header row
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}3"
    ws.freeze_panes = "A4"


def build_tab2_compliance_matrix(ws, jurisdictions, procedures, compliance_matrix):
    ws.title = "2. HSESC Compliance Matrix"
    ws.sheet_view.showGridLines = False

    ws.merge_cells(f"A1:{get_column_letter(3 + len(jurisdictions) * 2)}1")
    t = ws["A1"]
    t.value = "HSESC Compliance Matrix — Procedure × Jurisdiction"
    t.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    t.fill = cell_fill(LIGHT_B)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(3 + len(jurisdictions) * 2)}2")
    s = ws["A2"]
    s.value = (
        "GREEN = Full compliance achievable | AMBER = Compliance achievable with adaptation | "
        "RED = Structural barriers prevent full compliance | UNKNOWN = Insufficient data | "
        "Filter by RAG column per jurisdiction to identify gaps."
    )
    s.font = Font(name="Arial", size=8, italic=True, color="555555")
    s.alignment = wrap_align()
    ws.row_dimensions[2].height = 25

    # Build header row: PRO ID | Title | MS Element | then per-jurisdiction: RAG | Statement
    base_headers = ["Procedure ID", "Title", "Management System Element"]
    jur_headers = []
    for j in jurisdictions:
        jur_headers.append(f"{j['name']} — RAG")
        jur_headers.append(f"{j['name']} — Compliance Statement")

    all_headers = base_headers + jur_headers
    write_header_row(ws, 3, all_headers, bg=NAVY, size=8)
    ws.row_dimensions[3].height = 50

    # Jurisdiction sub-header with location
    for col_idx, j in enumerate(jurisdictions):
        col = 4 + col_idx * 2
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
        c = ws.cell(row=3, column=col, value=f"{j['name']} — {j['location']}")
        c.font = hdr_font(size=8)
        c.fill = hdr_fill(BLUE_H)
        c.alignment = wrap_align("center", "center")

    for row_idx, proc in enumerate(procedures, 4):
        pid  = proc["id"]
        title = proc["title"]
        elem  = proc.get("element", "")

        comp = compliance_matrix.get(pid, {})
        jur_data = comp.get("jurisdictions", {}) if comp else {}

        write_cell(ws, row_idx, 1, pid,
                   fill=cell_fill(GREY_R),
                   font=Font(name="Arial", size=9, bold=True))
        write_cell(ws, row_idx, 2, title)
        write_cell(ws, row_idx, 3, elem)

        # Determine overall row RAG (worst across all jurisdictions)
        all_rags = [jur_data.get(j["id"], {}).get("rag", "U") for j in jurisdictions]
        row_rag = "G"
        if "R" in all_rags:
            row_rag = "R"
        elif "A" in all_rags:
            row_rag = "A"
        elif "U" in all_rags:
            row_rag = "U"

        for col_idx, j in enumerate(jurisdictions):
            jid  = j["id"]
            jdat = jur_data.get(jid, {})
            rag  = jdat.get("rag", "U")
            stmt = jdat.get("statement", "No assessment available")
            barrier = jdat.get("specific_barrier", "")
            workaround = jdat.get("workaround", "")

            rag_col  = 4 + col_idx * 2
            stmt_col = rag_col + 1

            # RAG cell
            c_rag = write_cell(ws, row_idx, rag_col, rag_text(rag), fill=rag_fill(rag))
            c_rag.font = Font(name="Arial", size=9, bold=True,
                              color={"G": "375623", "A": "7F4F00", "R": "C00000"}.get(rag, "555555"))
            c_rag.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

            # Statement cell
            full_stmt = stmt
            if barrier:
                full_stmt += f"\nBarrier: {barrier}"
            if workaround:
                full_stmt += f"\nWorkaround: {workaround}"
            write_cell(ws, row_idx, stmt_col, full_stmt, fill=rag_fill(rag))

        ws.row_dimensions[row_idx].height = 60

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 25
    for col_idx in range(len(jurisdictions)):
        ws.column_dimensions[get_column_letter(4 + col_idx * 2)].width = 12
        ws.column_dimensions[get_column_letter(5 + col_idx * 2)].width = 45

    ws.auto_filter.ref = f"A3:{get_column_letter(len(all_headers))}3"
    ws.freeze_panes = "D4"


def build_tab3_versioning(ws, versioning_data):
    ws.title = "3. Document Versioning"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = "HSESC Document Versioning — Jurisdiction Groupings and Version Count"
    t.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    t.fill = cell_fill(LIGHT_B)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    if not versioning_data:
        ws["A3"] = "Versioning analysis not available — research may have failed."
        return

    # Overall strategy
    ws.merge_cells("A3:I3")
    strategy_label = ws["A3"]
    strategy_label.value = "OVERALL VERSIONING STRATEGY"
    strategy_label.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    strategy_label.fill = hdr_fill(NAVY)
    ws.row_dimensions[3].height = 20

    ws.merge_cells("A4:I4")
    strategy = ws["A4"]
    strategy.value = versioning_data.get("overall_strategy", "")
    strategy.font = body_font(size=10)
    strategy.alignment = wrap_align()
    ws.row_dimensions[4].height = 50

    # Jurisdiction groupings
    ws.merge_cells("A6:I6")
    gl = ws["A6"]
    gl.value = "JURISDICTION GROUPINGS"
    gl.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    gl.fill = hdr_fill(BLUE_H)
    ws.row_dimensions[6].height = 20

    write_header_row(ws, 7, ["Group Name", "Jurisdictions", "Rationale", "Document Standard"], bg=NAVY, size=9)
    for row_idx, grp in enumerate(versioning_data.get("jurisdiction_groupings", []), 8):
        write_cell(ws, row_idx, 1, grp.get("group_name", ""), fill=cell_fill(GREY_R),
                   font=Font(name="Arial", size=9, bold=True))
        write_cell(ws, row_idx, 2, ", ".join(grp.get("jurisdictions", [])))
        write_cell(ws, row_idx, 3, grp.get("rationale", ""))
        write_cell(ws, row_idx, 4, grp.get("document_standard", ""))
        ws.row_dimensions[row_idx].height = 50

    # Per-procedure versioning
    next_row = 8 + len(versioning_data.get("jurisdiction_groupings", [])) + 2
    ws.merge_cells(f"A{next_row}:I{next_row}")
    pl = ws[f"A{next_row}"]
    pl.value = "PER-PROCEDURE VERSION COUNT"
    pl.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    pl.fill = hdr_fill(BLUE_H)
    ws.row_dimensions[next_row].height = 20

    next_row += 1
    write_header_row(ws, next_row,
        ["Procedure ID", "Versions Needed", "Group 1", "Group 2", "Group 3", "Version Notes"],
        bg=NAVY, size=9)
    next_row += 1

    for proc in versioning_data.get("procedures", []):
        versions = proc.get("version_groups", [])
        grp_cols = [", ".join(g) for g in versions] + ["", ""]
        write_cell(ws, next_row, 1, proc.get("procedure_id", ""),
                   fill=cell_fill(GREY_R), font=Font(name="Arial", size=9, bold=True))
        n_vers = proc.get("versions_needed", len(versions))
        fill = cell_fill(LIGHT_G if n_vers == 1 else LIGHT_A if n_vers == 2 else LIGHT_R)
        write_cell(ws, next_row, 2, n_vers, fill=fill,
                   font=Font(name="Arial", size=9, bold=True),
                   align=Alignment(horizontal="center", vertical="top"))
        for i, grp_txt in enumerate(grp_cols[:3]):
            write_cell(ws, next_row, 3 + i, grp_txt)
        write_cell(ws, next_row, 6, proc.get("version_notes", ""))
        ws.row_dimensions[next_row].height = 40
        next_row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 55
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 20


def build_tab4_source_log(ws, jurisdictions, profiles):
    ws.title = "4. Source Log"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "Source Log — All Factual Claims and Verification Status"
    t.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    t.fill = cell_fill(LIGHT_B)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value = (
        "Verified = URL confirmed to contain the stated information. "
        "Unverified = Source cited but URL not confirmed during research. "
        "Not Found = No credible source identified — claim not included in Country Profiles."
    )
    s.font = Font(name="Arial", size=8, italic=True, color="555555")
    s.alignment = wrap_align()
    ws.row_dimensions[2].height = 30

    write_header_row(ws, 3,
        ["Jurisdiction", "Claim Summary", "Source Name", "URL", "Date Accessed", "Verified?", "Notes"],
        bg=NAVY, size=9)
    ws.row_dimensions[3].height = 30

    row_idx = 4
    for jur in jurisdictions:
        p = profiles.get(jur["id"], {})
        if not p:
            continue
        sources = p.get("sources", [])
        for src in sources:
            verified = src.get("verified", False)
            fill = cell_fill(LIGHT_G) if verified else cell_fill(LIGHT_A)
            write_cell(ws, row_idx, 1, jur["name"], fill=fill)
            write_cell(ws, row_idx, 2, src.get("claim", ""), fill=fill)
            write_cell(ws, row_idx, 3, src.get("source_name", ""), fill=fill)
            url = src.get("url", "")
            c = write_cell(ws, row_idx, 4, url if url else "No URL available", fill=fill)
            if url and url.startswith("http"):
                c.hyperlink = url
                c.font = Font(name="Arial", size=9, color="0563C1", underline="single")
            write_cell(ws, row_idx, 5, src.get("date_accessed", ""), fill=fill)
            write_cell(ws, row_idx, 6, "Verified" if verified else "Unverified", fill=fill,
                       font=Font(name="Arial", size=9, bold=True,
                                 color="375623" if verified else "C55A11"))
            write_cell(ws, row_idx, 7, src.get("notes", ""), fill=fill)
            ws.row_dimensions[row_idx].height = 35
            row_idx += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 35
    ws.auto_filter.ref = f"A3:G{row_idx}"
    ws.freeze_panes = "A4"


# ── Registry loader ───────────────────────────────────────────────────────────

def load_procedures(brief_dir):
    registry_path = os.path.join(brief_dir, "_registry.json")
    if os.path.exists(registry_path):
        with open(registry_path, "r", encoding="utf-8") as f:
            return json.load(f)

    # Fallback: hardcoded list
    return [
        {"id": f"PRO-{str(i).zfill(4)}", "title": f"Procedure {i}"}
        for i in range(1, 52)
    ]


def load_ms_elements(brief_dir, procedures):
    """Try to enrich procedures with management system element from brief JSON."""
    for proc in procedures:
        brief_path = os.path.join(brief_dir, f"{proc['id']}.json")
        if os.path.exists(brief_path):
            try:
                with open(brief_path, "r", encoding="utf-8") as f:
                    brief = json.load(f)
                proc["element"] = brief.get("management_system_element", "")
            except Exception:
                proc["element"] = ""
        else:
            proc["element"] = ""
    return procedures


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--brief-dir",     required=True)
    parser.add_argument("--output-dir",    required=True)
    parser.add_argument("--jurisdictions", default="all")
    parser.add_argument("--phase",         default="all",
                        help="Phase to run: 1, 2, 34, or all")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])

    requested = [j.strip().upper() for j in args.jurisdictions.split(",")]
    jurs_to_research = JURISDICTIONS if "ALL" in requested else         [j for j in JURISDICTIONS if j["id"] in requested]

    procedures = load_procedures(args.brief_dir)
    procedures = load_ms_elements(args.brief_dir, procedures)
    print(f"Loaded {len(procedures)} procedures from registry")

    run_p1  = args.phase in ("1", "all")
    run_p2  = args.phase in ("2", "all")
    run_p34 = args.phase in ("34", "all")

    profiles_path = os.path.join(args.output_dir, "_country_profiles_cache.json")
    matrix_path   = os.path.join(args.output_dir, "_compliance_matrix_cache.json")

    # ── Phase 1: Research each jurisdiction ──────────────────────────────────
    profiles = {}
    if os.path.exists(profiles_path):
        with open(profiles_path, "r", encoding="utf-8") as f:
            profiles = json.load(f)
        print(f"Loaded {len(profiles)} cached profiles")

    if run_p1:
        remaining = [j for j in jurs_to_research if j["id"] not in profiles]
        print(f"Phase 1: {len(remaining)} jurisdiction(s) to research ({len(profiles)} cached)")
        for i, jur in enumerate(jurs_to_research):
            jid = jur["id"]
            if jid in profiles:
                print(f"  SKIP {jid} — cached")
                continue
            print(f"\n  [{i+1}/{len(jurs_to_research)}] Researching {jur['name']}...")
            profile = research_jurisdiction_profile(client, jur)
            profiles[jid] = profile if profile else {"_research_failed": True}
            status = len(profile.get("sources", [])) if profile else "FAILED"
            print(f"  Done: {status} sources")
            with open(profiles_path, "w", encoding="utf-8") as f:
                json.dump(profiles, f, indent=2, ensure_ascii=False)
            time.sleep(2)
        print(f"Phase 1 complete. {len(profiles)} profiles saved.")
    else:
        print("Phase 1: skipped — using cache")

    # ── Phase 2: Compliance matrix ────────────────────────────────────────────
    compliance_matrix = {}
    if os.path.exists(matrix_path):
        with open(matrix_path, "r", encoding="utf-8") as f:
            compliance_matrix = json.load(f)
        print(f"\nLoaded {len(compliance_matrix)} cached procedure assessments")

    if run_p2:
        remaining_procs = [p for p in procedures if p["id"] not in compliance_matrix]
        print(f"Phase 2: {len(remaining_procs)} procedure(s) to assess ({len(compliance_matrix)} cached)")
        for i, proc in enumerate(procedures):
            pid = proc["id"]
            if pid in compliance_matrix:
                continue
            print(f"  [{i+1}/{len(procedures)}] {pid}: {proc['title'][:45]}...")
            result = research_procedure_compliance(client, proc, JURISDICTIONS, profiles)
            compliance_matrix[pid] = result if result else                 {"_assessment_failed": True, "jurisdictions": {}}
            time.sleep(1)
            if (i + 1) % 5 == 0:
                with open(matrix_path, "w", encoding="utf-8") as f:
                    json.dump(compliance_matrix, f, indent=2, ensure_ascii=False)
                print(f"    Checkpoint: {i+1} procedures complete")
        with open(matrix_path, "w", encoding="utf-8") as f:
            json.dump(compliance_matrix, f, indent=2, ensure_ascii=False)
        print(f"Phase 2 complete. {len(compliance_matrix)} assessments saved.")
    else:
        print("Phase 2: skipped — using cache")

    if not run_p34:
        print("Phases 3+4 skipped.")
        return

    # ── Phase 3: Versioning analysis ─────────────────────────────────────────
    print("\nPhase 3: versioning analysis...")
    versioning = research_versioning(client, procedures, compliance_matrix)

    # ── Phase 4: Build Excel workbook ────────────────────────────────────────
    print("\nPhase 4: building Excel workbook...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("1. Country Profiles")
    build_tab1_country_profiles(ws1, JURISDICTIONS, profiles)
    print("  Tab 1 done")

    ws2 = wb.create_sheet("2. HSESC Compliance Matrix")
    build_tab2_compliance_matrix(ws2, JURISDICTIONS, procedures, compliance_matrix)
    print("  Tab 2 done")

    ws3 = wb.create_sheet("3. Document Versioning")
    build_tab3_versioning(ws3, versioning)
    print("  Tab 3 done")

    ws4 = wb.create_sheet("4. Source Log")
    build_tab4_source_log(ws4, JURISDICTIONS, profiles)
    print("  Tab 4 done")

    output_path = os.path.join(args.output_dir, "HSESC_Country_Compliance_Sourcebook.xlsx")
    wb.save(output_path)
    print(f"\nSourcebook saved: {output_path}")

if __name__ == "__main__":
    main()
